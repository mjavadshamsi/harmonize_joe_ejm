import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# -----------------------------------------------------
# Configuration and Environment
# -----------------------------------------------------
# Adjust or remove directory-changing logic as needed
if os.environ.get('USER') == 'XXX':
    os.chdir('x_directory')
elif os.environ.get('USERNAME') == 'YYY':
    os.chdir('Y_directory')
else:
    print("working directory was not changed.")

# Excluded countries
excluded_countries = [
    'china', 'taiwan', 'philippines', 'hong', 'malawi', 'colombia', 'japan',
    'brazil', 'chile', 'india', 'israel', 'argentina', 'mexico', 'saudi arabia',
    'turkey', 'south africa', 'bangladesh', 'lebanon', 'uzbekistan', 'russia',
    'korea', 'saudi'
]

# Master columns
master_columns = [
    'institution', 'division', 'department', 'keywords', 'title',
    'deadline', 'country', 'jp_id', 'ejm_id', 'joe_url',
    'ejm_url', 'BatchDate', 'Source'
]

# -----------------------------------------------------
# Utility Functions
# -----------------------------------------------------
def reorder_and_fill_columns(df, column_order=master_columns):
    """
    Ensure df has all columns in column_order, filling missing with None,
    and returns df in that exact order.
    """
    for col in column_order:
        if col not in df.columns:
            df[col] = None
    return df[column_order]

def append_df_to_ws(ws, df):
    """
    Append rows of a DataFrame to an openpyxl worksheet without overwriting existing data.
    """
    if df.empty:
        return
    rows = dataframe_to_rows(df, index=False, header=False)
    for row in rows:
        ws.append(row)

def get_latest_file(directory, pattern):
    """
    Returns the path of the latest file (by date in the filename) from the given directory,
    matching the given regex pattern.
    """
    files = [f for f in os.listdir(directory) if re.match(pattern, f)]
    if not files:
        return None

    # Parse dates from filenames
    compiled_pattern = re.compile(pattern)
    file_dates = []
    for f in files:
        match = compiled_pattern.search(f)
        if match:
            # Expecting group(1)=day, group(2)=month, group(3)=year
            day = int(match.group(1))
            month = int(match.group(2))
            year = int(match.group(3))
            date_obj = datetime(year, month, day)
            file_dates.append((date_obj, f))

    if not file_dates:
        return None

    # Sort descending by date and return latest
    file_dates.sort(reverse=True, key=lambda x: x[0])
    return os.path.join(directory, file_dates[0][1])

def load_or_create_master(file_name):
    """
    Loads the Excel master file if it exists; otherwise, create it with empty 'Listings' and 'Deleted' sheets.
    Returns (workbook, listings_df, deleted_df).
    """
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        if 'Listings' not in wb.sheetnames:
            wb.create_sheet('Listings')
        if 'Deleted' not in wb.sheetnames:
            wb.create_sheet('Deleted')

        df_listings = pd.read_excel(file_name, sheet_name='Listings') if len(wb['Listings'].values) > 1 else pd.DataFrame()
        df_deleted = pd.read_excel(file_name, sheet_name='Deleted') if len(wb['Deleted'].values) > 1 else pd.DataFrame()
    else:
        wb = Workbook()
        # Remove the default sheet created by openpyxl
        default_sheet = wb.active
        wb.remove(default_sheet)
        wb.create_sheet('Listings')
        wb.create_sheet('Deleted')
        df_listings = pd.DataFrame()
        df_deleted = pd.DataFrame()
        wb.save(file_name)

    return wb, df_listings.convert_dtypes(), df_deleted.convert_dtypes()

def save_master(file_name, wb):
    """Saves the workbook to the specified file name."""
    wb.save(file_name)

# -----------------------------------------------------
# Process JOE Listings
# -----------------------------------------------------
# 1. Identify the latest joe_resultset file
joe_pattern = r'joe_resultset_(\d{2})_(\d{2})_(\d{4})'
latest_joe_file = get_latest_file('joe_listings', joe_pattern)
if latest_joe_file:
    print(f"[JOE] Latest file: {latest_joe_file}")
    # 2. Read & harmonize
    df_joe = pd.read_excel(latest_joe_file)

    # Extract country from 'locations' field (custom logic)
    def get_country(location):
        if pd.isna(location):
            return None
        words = location.split()
        for word in words:
            if word.isupper():
                return word
        return None

    df_joe['country'] = df_joe['locations'].apply(get_country)

    # Remove unwanted columns
    remove_cols = [
        'joe_issue_ID', 'jp_section', 'jp_full_text', 'jp_agency_insertion_num',
        'locations', 'JEL_Classifications', 'salary_range', 'Date_Active'
    ]
    df_joe = df_joe.drop(columns=[c for c in remove_cols if c in df_joe.columns], errors='ignore')

    # Rename columns
    new_cols = []
    for col in df_joe.columns:
        if col == 'jp_id':
            new_cols.append(col)
        elif col.startswith('jp_'):
            new_cols.append(col.replace('jp_', ''))
        else:
            new_cols.append(col)
    df_joe.columns = new_cols
    df_joe = df_joe.rename(columns={'Application_deadline': 'deadline'})

    # Convert deadline to date
    if 'deadline' in df_joe.columns:
        df_joe['deadline'] = pd.to_datetime(df_joe['deadline'], errors='coerce').dt.date

    # Create joe_url
    if 'jp_id' in df_joe.columns:
        df_joe['joe_url'] = df_joe['jp_id'].apply(
            lambda x: f"https://www.aeaweb.org/joe/listing.php?JOE_ID={x}" if pd.notna(x) else None
        )

    # Reorder/fill columns
    df_joe = reorder_and_fill_columns(df_joe)

    # Extract batch date from filename
    match = re.search(joe_pattern, os.path.basename(latest_joe_file))
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        batch_date_str = f"J_{year}-{month}-{day}"
    else:
        batch_date_str = "J_unknown"

    df_joe['BatchDate'] = batch_date_str
    df_joe['Source'] = "JOE"

    # 3. Load or create master
    excel_file = 'Application_MasterFile.xlsx'
    wb, df_listings, df_deleted = load_or_create_master(excel_file)

    # 4. Existing IDs
    existing_ids = set()
    if not df_listings.empty:
        if 'jp_id' in df_listings.columns:
            existing_ids.update(df_listings['jp_id'].dropna().tolist())
        if 'ejm_id' in df_listings.columns:
            existing_ids.update(df_listings['ejm_id'].dropna().tolist())
    if not df_deleted.empty:
        if 'jp_id' in df_deleted.columns:
            existing_ids.update(df_deleted['jp_id'].dropna().tolist())
        if 'ejm_id' in df_deleted.columns:
            existing_ids.update(df_deleted['ejm_id'].dropna().tolist())

    # 5. Filter new rows
    df_joe_new = df_joe[~df_joe['jp_id'].isin(existing_ids)].copy()

    # Combine current master
    df_master = pd.concat([df_listings, df_deleted], ignore_index=True)
    if df_master.empty or 'BatchDate' not in df_master.columns:
        df_master['BatchDate'] = pd.NaT

    # Decide listing vs deleted
    listings_to_add = []
    deleted_to_add = []

    for idx, row in df_joe_new.iterrows():
        this_batch = row['BatchDate']
        country = row['country'].lower() if pd.notna(row['country']) else ''
        # Check if any row with this batch is already in master
        batch_in_master = df_master[df_master['BatchDate'] == this_batch]

        # If country is excluded OR batch_in_master not empty -> Deleted
        if any(ec in country for ec in excluded_countries) or (not batch_in_master.empty):
            deleted_to_add.append(row)
        else:
            listings_to_add.append(row)

    df_listings_add = pd.DataFrame(listings_to_add)
    df_deleted_add = pd.DataFrame(deleted_to_add)

    # 6. Append data to workbook
    ws_listings = wb['Listings']
    ws_deleted = wb['Deleted']
    append_df_to_ws(ws_listings, df_listings_add)
    append_df_to_ws(ws_deleted, df_deleted_add)

    # 7. Save
    save_master(excel_file, wb)
    print("[JOE] Applications file updated successfully.")
else:
    print("[JOE] No latest file found. Skipping JOE processing.")

# -----------------------------------------------------
# Process EJM Listings
# -----------------------------------------------------
ejm_pattern = r'positions_(\d{2})_(\d{2})_(\d{4})'
latest_ejm_file = get_latest_file('ejm_listings', ejm_pattern)
if latest_ejm_file:
    print(f"[EJM] Latest file: {latest_ejm_file}")

    # For EJM, we assume CSV; adjust if XLSX
    # 1. Read columns
    temp_columns = pd.read_csv(latest_ejm_file, header=None, skiprows=1, nrows=1).values.flatten().tolist()
    # 2. Read the data
    df_ejm = pd.read_csv(latest_ejm_file, skiprows=2, names=temp_columns)

    # Rename columns for consistency
    df_ejm = df_ejm.rename(columns={
        'Id': 'ejm_id',
        'URL': 'ejm_url',
        'Ad title': 'title',
        'Types': 'section',
        'Categories': 'keywords',
        'Deadline': 'deadline',
        'Department': 'department',
        'Institution': 'institution',
        'Country': 'country',
        'Application method': 'application_method'
    })

    # Keep only relevant columns
    keep_cols = ['ejm_id', 'ejm_url', 'title', 'keywords', 'deadline',
                 'department', 'institution', 'country']
    df_ejm = df_ejm[keep_cols]

    # Convert deadline
    df_ejm['deadline'] = pd.to_datetime(df_ejm['deadline'], errors='coerce').dt.date

    # Add missing JP columns for consistency
    df_ejm['jp_id'] = None
    df_ejm['joe_url'] = None  # or create if needed

    # Reorder
    df_ejm = reorder_and_fill_columns(df_ejm)

    # Batch date from filename
    match = re.search(ejm_pattern, os.path.basename(latest_ejm_file))
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        batch_date_str = f"E_{year}-{month}-{day}"
    else:
        batch_date_str = "E_unknown"

    df_ejm['BatchDate'] = batch_date_str
    df_ejm['Source'] = "EJM"

    # 3. Load or create master (re-load for EJM step)
    excel_file = 'Application_MasterFile.xlsx'
    wb, df_listings, df_deleted = load_or_create_master(excel_file)

    # 4. Existing IDs
    existing_ids = set()
    if not df_listings.empty:
        if 'jp_id' in df_listings.columns:
            existing_ids.update(df_listings['jp_id'].dropna().tolist())
        if 'ejm_id' in df_listings.columns:
            existing_ids.update(df_listings['ejm_id'].dropna().tolist())
    if not df_deleted.empty:
        if 'jp_id' in df_deleted.columns:
            existing_ids.update(df_deleted['jp_id'].dropna().tolist())
        if 'ejm_id' in df_deleted.columns:
            existing_ids.update(df_deleted['ejm_id'].dropna().tolist())

    # 5. Filter new EJM rows
    df_ejm_new = df_ejm[~df_ejm['ejm_id'].isin(existing_ids)].copy()

    df_master = pd.concat([df_listings, df_deleted], ignore_index=True)
    if df_master.empty or 'BatchDate' not in df_master.columns:
        df_master['BatchDate'] = pd.NaT

    listings_to_add = []
    deleted_to_add = []

    for idx, row in df_ejm_new.iterrows():
        this_batch = row['BatchDate']
        country = row['country'].lower() if pd.notna(row['country']) else ''
        batch_in_master = df_master[df_master['BatchDate'] == this_batch]

        # If country is excluded or batch_in_master not empty -> Deleted
        if not country or any(ec in country for ec in excluded_countries) or (not batch_in_master.empty):
            deleted_to_add.append(row)
        else:
            listings_to_add.append(row)

    df_listings_add = pd.DataFrame(listings_to_add)
    df_deleted_add = pd.DataFrame(deleted_to_add)

    # 6. Append
    ws_listings = wb['Listings']
    ws_deleted = wb['Deleted']
    append_df_to_ws(ws_listings, df_listings_add)
    append_df_to_ws(ws_deleted, df_deleted_add)

    # 7. Save
    save_master(excel_file, wb)
    print("[EJM] Applications file updated successfully.")
else:
    print("[EJM] No latest file found. Skipping EJM processing.")
